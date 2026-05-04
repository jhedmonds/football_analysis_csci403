from pathlib import Path
import xml.etree.ElementTree as ET

import re
import pdfplumber
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


# ------------------------------------------------------------
# Project folders
# ------------------------------------------------------------
# Path(__file__).parent means "the folder this scraper.py file is in"
BASE_DIR = Path(__file__).parent

INPUT_DIR = BASE_DIR / "data" / "input_files"
EXCEL_DIR = BASE_DIR / "data" / "excel_files"
OUTPUT_DIR = BASE_DIR / "output"

OUTPUT_DIR.mkdir(exist_ok=True)


# ------------------------------------------------------------
# Excel files for each team
# ------------------------------------------------------------
TEAM_FILES = {
    "csup": "csup_football_stats.xlsx",
    "mesa": "mesa_football_stats.xlsx",
    "mines": "mines_football_stats.xlsx",
    "western": "western_football_stats.xlsx",
}


# ------------------------------------------------------------
# Maps Excel column names to possible names in XML/HTML files
# ------------------------------------------------------------
# Left side = exact column name in Excel
# Right side = possible stat names from scraped files
STAT_ALIASES = {
    "Rush Yards": ["rush_yards", "rushing_yards"],
    "Rush Yards Gained": ["rush_yards_gained", "rushing_yards_gained"],
    "Rush Yards Lost": ["rush_yards_lost", "rushing_yards_lost"],
    "Rush Attempts": ["rush_attempts", "rushing_attempts", "carries"],
    "Avg Rush Yards/Attempt": ["avg_rush_yards_attempt", "yards_per_rush"],
    "Avg Rush Yards/Game": ["avg_rush_yards_game"],
    "Rush TDs": ["rush_tds", "rushing_tds"],

    "Pass Attempts": ["pass_attempts", "passing_attempts"],
    "Avg Pass Yards/Attempt": ["avg_pass_yards_attempt", "yards_per_pass"],
    "Avg Pass Yards/Game": ["avg_pass_yards_game"],
    "Pass TDs": ["pass_tds", "passing_tds"],

    "Total Offensive Plays": ["total_offensive_plays", "offensive_plays"],
    "Avg Yards/Play": ["avg_yards_play", "yards_per_play"],
    "Avg Yards/Game": ["avg_yards_game"],
    "Interceptions Thrown": ["interceptions_thrown", "ints_thrown"],
    "3rd Down Conv.": ["third_down_conv", "third_down_conversion", "3rd_down_conv"],
    "4th Down Conv.": ["fourth_down_conv", "fourth_down_conversion", "4th_down_conv"],
    "Sacks Taken": ["sacks_taken"],
    "TDs Scored": ["tds_scored", "touchdowns"],
    "Time of Possession/Game (min)": ["time_of_possession_game", "top_game"],
    "Total Penalties": ["total_penalties", "penalties"],
    "Avg Penalty Yards/Game": ["avg_penalty_yards_game"],

    "Opp. Rush Yards": ["opp_rush_yards", "opponent_rush_yards"],
    "Opp. Rush Yards Gained": ["opp_rush_yards_gained"],
    "Opp. Rush Yards Lost": ["opp_rush_yards_lost"],
    "Opp. Rush Attempts": ["opp_rush_attempts"],
    "Opp. Avg Rush Yards/Attempt": ["opp_avg_rush_yards_attempt"],
    "Opp. Avg Rush Yards/Game": ["opp_avg_rush_yards_game"],
    "Opp. Rush TDs": ["opp_rush_tds"],

    "Opp. Pass Attempts": ["opp_pass_attempts"],
    "Opp. Avg Pass Yards/Attempt": ["opp_avg_pass_yards_attempt"],
    "Opp. Avg Pass Yards/Game": ["opp_avg_pass_yards_game"],
    "Opp. Pass TDs": ["opp_pass_tds"],

    "Total Defensive Plays": ["total_defensive_plays", "defensive_plays"],
    "Opp. Avg Yards/Play": ["opp_avg_yards_play"],
    "Opp. Avg Yards/Game": ["opp_avg_yards_game"],
    "Interceptions": ["interceptions", "ints"],
    "Opp. 3rd Down Conv.": ["opp_third_down_conv"],
    "Opp. 4th Down Conv.": ["opp_fourth_down_conv"],
    "Sacks": ["sacks"],
    "Opp. TDs Scored": ["opp_tds_scored"],
    "Opp. Time of Possession/Game (min)": ["opp_time_of_possession_game"],
    "Opp. Total Penalties": ["opp_total_penalties"],
    "Opp. Avg Penalty Yards/Game": ["opp_avg_penalty_yards_game"],
}


def get_two_values_after_label(line, label):
    """
    Finds the first two values immediately after a stat label.

    Example:
    "Yards gained 2404 1620 *10/14/2017 ..."
    returns:
    ("2404", "1620")
    """
    pattern = rf"^{re.escape(label)}\s+([0-9.,:-]+)\s+([0-9.,:-]+)"
    match = re.search(pattern, line.strip(), re.IGNORECASE)

    if not match:
        return None, None

    return match.group(1), match.group(2)


def get_two_dash_values_after_label(line, label):
    """
    Finds first two dash-formatted values after a label.

    Example:
    "Att-Comp-Int 349-209-10 380-186-23"
    returns:
    ("349-209-10", "380-186-23")
    """
    pattern = rf"^{re.escape(label)}\s+([0-9]+-[0-9]+-[0-9]+)\s+([0-9]+-[0-9]+-[0-9]+)"
    match = re.search(pattern, line.strip(), re.IGNORECASE)

    if not match:
        return None, None

    return match.group(1), match.group(2)


def get_two_conversion_values_after_label(line, label):
    """
    Finds first two conversion values after a label.

    Example:
    "3RD-DOWN CONVERSIONS 67-155 57-183"
    returns:
    ("67-155", "57-183")
    """
    pattern = rf"^{re.escape(label)}\s+([0-9]+-[0-9]+)\s+([0-9]+-[0-9]+)"
    match = re.search(pattern, line.strip(), re.IGNORECASE)

    if not match:
        return None, None

    return match.group(1), match.group(2)


def time_to_minutes(time_text):
    """
    Converts time like 29:28 into decimal minutes.

    Example:
    29:28 -> 29.47
    """
    if ":" not in str(time_text):
        return clean_value(time_text)

    minutes, seconds = str(time_text).split(":")
    return round(int(minutes) + int(seconds) / 60, 2)


def clean_key(text):
    """
    Standardizes stat names so different formats can still match.

    Example:
    "Rush Yards" -> "rush_yards"
    "Rush.Yards" -> "rushyards"
    """
    return (
        str(text)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace(".", "")
        .replace("/", "_")
        .replace("-", "_")
    )


def clean_value(value):
    """
    Converts scraped text values into useful Python values.

    Examples:
    "1850" -> 1850
    "4.5" -> 4.5
    "55%" -> 0.55
    """
    if value is None:
        return None

    value = str(value).strip()

    if value == "":
        return None

    if value.endswith("%"):
        return float(value.replace("%", "")) / 100

    try:
        return int(value)
    except ValueError:
        pass

    try:
        return float(value)
    except ValueError:
        return value
    

def conversion_to_percentage(value):
    """
    Converts "67-155" → 0.4323

    Returns a float (0–1 range) so Excel can format as percentage.
    """
    if not value or "-" not in str(value):
        return None

    try:
        made, attempts = value.split("-")
        made = float(made)
        attempts = float(attempts)

        if attempts == 0:
            return 0

        return round(made / attempts, 4)
    except:
        return None


def parse_xml_file(file_path):
    """
    Reads an XML file and stores each tag/value pair in a dictionary.

    Example XML:
    <rush_yards>1850</rush_yards>

    Becomes:
    {"rush_yards": 1850}
    """
    tree = ET.parse(file_path)
    root = tree.getroot()

    stats = {}

    for elem in root.iter():
        key = clean_key(elem.tag)
        value = clean_value(elem.text)

        if value is not None:
            stats[key] = value

        # Also checks XML attributes, just in case the data is stored there.
        for attr_key, attr_value in elem.attrib.items():
            stats[clean_key(attr_key)] = clean_value(attr_value)

    return stats


def parse_html_file(file_path):
    """
    Reads an HTML file and tries to pull stats from table rows.

    Expected simple table format:
    <tr><th>rush_yards</th><td>1850</td></tr>
    """
    html = file_path.read_text(encoding="utf-8", errors="ignore")
    soup = BeautifulSoup(html, "lxml")

    stats = {}

    for table in soup.find_all("table"):
        for row in table.find_all("tr"):
            cells = row.find_all(["th", "td"])
            cell_text = [cell.get_text(strip=True) for cell in cells]

            # This assumes the first cell is the stat name
            # and the second cell is the stat value.
            if len(cell_text) >= 2:
                key = clean_key(cell_text[0])
                value = clean_value(cell_text[1])

                if value is not None:
                    stats[key] = value

    return stats


def parse_stat_lines(lines):
    """
    Looks through text lines and tries to find stat/value pairs.
    This works for text extracted from PDFs or webpages.
    """
    stats = {}

    alias_names = []

    for excel_header, aliases in STAT_ALIASES.items():
        alias_names.append(excel_header)

        for alias in aliases:
            alias_names.append(alias)

    for line in lines:
        cleaned_line = line.strip()

        if not cleaned_line:
            continue

        lower_line = cleaned_line.lower()

        for alias in alias_names:
            readable_alias = alias.replace("_", " ")
            lower_alias = readable_alias.lower()

            if lower_line.startswith(lower_alias):
                value_text = cleaned_line[len(readable_alias):].strip()
                value = clean_value(value_text)

                if value is not None:
                    stats[clean_key(alias)] = value

    return stats





def parse_pdf_file(file_path):
    """
    Reads the football PDF and extracts only the first-page Team Statistics
    section. This avoids pulling values from player tables or game schedule rows.
    """
    stats = {}

    with pdfplumber.open(file_path) as pdf:
        first_page_text = pdf.pages[0].extract_text()

    if not first_page_text:
        return stats

    lines = first_page_text.splitlines()
    current_section = None

    for line in lines:
        line = line.strip()

        if not line:
            continue

        # Track which section we are currently in.
        if line.startswith("RUSHING YARDAGE"):
            current_section = "rushing"
            team_value, opp_value = get_two_values_after_label(line, "RUSHING YARDAGE")
            stats["rush_yards"] = clean_value(team_value)
            stats["opp_rush_yards"] = clean_value(opp_value)

        elif line.startswith("PASSING YARDAGE"):
            current_section = "passing"
            team_value, opp_value = get_two_values_after_label(line, "PASSING YARDAGE")
            stats["pass_yards"] = clean_value(team_value)
            stats["opp_pass_yards"] = clean_value(opp_value)

        elif line.startswith("TOTAL OFFENSE"):
            current_section = "total_offense"
            team_value, opp_value = get_two_values_after_label(line, "TOTAL OFFENSE")
            stats["total_offense"] = clean_value(team_value)
            stats["opp_total_offense"] = clean_value(opp_value)

        elif line.startswith("PENALTIES-YARDS"):
            current_section = "penalties"
            team_value, opp_value = get_two_conversion_values_after_label(line, "PENALTIES-YARDS")

            if team_value and opp_value:
                stats["total_penalties"] = clean_value(team_value.split("-")[0])
                stats["opp_total_penalties"] = clean_value(opp_value.split("-")[0])

        elif line.startswith("TIME OF POSSESSION"):
            current_section = "time_of_possession"

        # Rushing stats
        elif line.startswith("Yards gained") and current_section == "rushing":
            team_value, opp_value = get_two_values_after_label(line, "Yards gained")
            stats["rush_yards_gained"] = clean_value(team_value)
            stats["opp_rush_yards_gained"] = clean_value(opp_value)

        elif line.startswith("Yards lost") and current_section == "rushing":
            team_value, opp_value = get_two_values_after_label(line, "Yards lost")
            stats["rush_yards_lost"] = clean_value(team_value)
            stats["opp_rush_yards_lost"] = clean_value(opp_value)

        elif line.startswith("Attempts") and current_section == "rushing":
            team_value, opp_value = get_two_values_after_label(line, "Attempts")
            stats["rush_attempts"] = clean_value(team_value)
            stats["opp_rush_attempts"] = clean_value(opp_value)

        elif line.startswith("TDs Passing") and current_section == "rushing":
            team_value, opp_value = get_two_values_after_label(line, "TDs Passing")
            stats["rush_tds"] = clean_value(team_value)
            stats["opp_rush_tds"] = clean_value(opp_value)

        # Passing stats
        elif line.startswith("Att-Comp-Int") and current_section == "passing":
            team_value, opp_value = get_two_dash_values_after_label(line, "Att-Comp-Int")

            if team_value and opp_value:
                stats["pass_attempts"] = clean_value(team_value.split("-")[0])
                stats["opp_pass_attempts"] = clean_value(opp_value.split("-")[0])
                stats["interceptions_thrown"] = clean_value(team_value.split("-")[2])
                stats["interceptions"] = clean_value(opp_value.split("-")[2])

        elif line.startswith("Touchdowns") and current_section == "passing":
            team_value, opp_value = get_two_values_after_label(line, "Touchdowns")
            stats["pass_tds"] = clean_value(team_value)
            stats["opp_pass_tds"] = clean_value(opp_value)

        # Total offense stats
        elif line.startswith("Total Plays") and current_section == "total_offense":
            team_value, opp_value = get_two_values_after_label(line, "Total Plays")
            stats["total_offensive_plays"] = clean_value(team_value)
            stats["total_defensive_plays"] = clean_value(opp_value)

        elif line.startswith("Avg Per Play") and current_section == "total_offense":
            team_value, opp_value = get_two_values_after_label(line, "Avg Per Play")
            stats["avg_yards_play"] = clean_value(team_value)
            stats["opp_avg_yards_play"] = clean_value(opp_value)

        # Repeated "Avg Per Attempt" and "Avg Per Game" lines depend on section.
        elif line.startswith("Avg Per Attempt"):
            team_value, opp_value = get_two_values_after_label(line, "Avg Per Attempt")

            if current_section == "rushing":
                stats["avg_rush_yards_attempt"] = clean_value(team_value)
                stats["opp_avg_rush_yards_attempt"] = clean_value(opp_value)

            elif current_section == "passing":
                stats["avg_pass_yards_attempt"] = clean_value(team_value)
                stats["opp_avg_pass_yards_attempt"] = clean_value(opp_value)

        elif line.startswith("Avg Per Game"):
            team_value, opp_value = get_two_values_after_label(line, "Avg Per Game")

            if current_section == "rushing":
                stats["avg_rush_yards_game"] = clean_value(team_value)
                stats["opp_avg_rush_yards_game"] = clean_value(opp_value)

            elif current_section == "passing":
                stats["avg_pass_yards_game"] = clean_value(team_value)
                stats["opp_avg_pass_yards_game"] = clean_value(opp_value)

            elif current_section == "total_offense":
                stats["avg_yards_game"] = clean_value(team_value)
                stats["opp_avg_yards_game"] = clean_value(opp_value)

            elif current_section == "penalties":
                stats["avg_penalty_yards_game"] = clean_value(team_value)
                stats["opp_avg_penalty_yards_game"] = clean_value(opp_value)

        elif line.startswith("Per Game") and current_section == "time_of_possession":
            team_value, opp_value = get_two_values_after_label(line, "Per Game")
            stats["time_of_possession_game"] = time_to_minutes(team_value)
            stats["opp_time_of_possession_game"] = time_to_minutes(opp_value)

        # Conversions
        elif line.startswith("3RD-DOWN CONVERSIONS"):
            team_value, opp_value = get_two_conversion_values_after_label(line, "3RD-DOWN CONVERSIONS")
            
            stats["third_down_conv"] = conversion_to_percentage(team_value)
            stats["opp_third_down_conv"] = conversion_to_percentage(opp_value)

        elif line.startswith("4TH-DOWN CONVERSIONS"):
            team_value, opp_value = get_two_conversion_values_after_label(line, "4TH-DOWN CONVERSIONS")
            
            stats["fourth_down_conv"] = conversion_to_percentage(team_value)
            stats["opp_fourth_down_conv"] = conversion_to_percentage(opp_value)

        # Sacks
        elif line.startswith("SACKS BY-YARDS"):
            team_value, opp_value = get_two_conversion_values_after_label(line, "SACKS BY-YARDS")

            if team_value and opp_value:
                stats["sacks"] = clean_value(team_value.split("-")[0])
                stats["sacks_taken"] = clean_value(opp_value.split("-")[0])

        # Touchdowns
        elif line.startswith("TOUCHDOWNS SCORED"):
            team_value, opp_value = get_two_values_after_label(line, "TOUCHDOWNS SCORED")
            stats["tds_scored"] = clean_value(team_value)
            stats["opp_tds_scored"] = clean_value(opp_value)

    return stats


def parse_url_file(url):
    """
    Parses the CSU Pueblo clean=true stats webpage using section-based regex.
    This works even when the HTML table structure is weird.
    """
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0 Safari/537.36"
        )
    }

    response = requests.get(url, headers=headers, timeout=15)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "lxml")
    text = soup.get_text(" ", strip=True)
    text = re.sub(r"\s+", " ", text)

    stats = {}

    def get_section(start_label, end_label):
        # First try exact case. This catches real section headers like RUSHING/PASSING.
        match = re.search(
            rf"{re.escape(start_label)}(.*?){re.escape(end_label)}",
            text
        )

        if match:
            return match.group(1)

        # Fallback for tests or pages where capitalization changes.
        match = re.search(
            rf"{re.escape(start_label)}(.*?){re.escape(end_label)}",
            text,
            re.IGNORECASE
        )

        return match.group(1) if match else ""

    def two_numbers(section, pattern):
        match = re.search(pattern, section, re.IGNORECASE)
        if not match:
            return None, None
        return match.group(1), match.group(2)

    def two_ratios(section, pattern):
        match = re.search(pattern, section, re.IGNORECASE)
        if not match:
            return None, None
        return match.group(1), match.group(2)

    def get_section_by_pattern(start_pattern, end_pattern):
        match = re.search(
            rf"{start_pattern}(.*?){end_pattern}",
            text,
            re.IGNORECASE
        )

        return match.group(1) if match else ""


    rushing = get_section_by_pattern(
        r"RUSHING\s+Yards Gained",
        r"PASSING\s+Att-Comp-Int"
    )

    # Add the first label back because the section starts after "Yards Gained"
    rushing = "Yards Gained " + rushing if rushing else ""

    passing = get_section_by_pattern(
        r"PASSING\s+Att-Comp-Int",
        r"TOTAL OFFENSE\s+Total Plays"
    )

    passing = "Att-Comp-Int " + passing if passing else ""

    total_offense = get_section_by_pattern(
        r"TOTAL OFFENSE\s+Total Plays",
        r"RETURNS\s+Kickoff"
    )

    total_offense = "Total Plays " + total_offense if total_offense else ""

    penalties = get_section_by_pattern(
        r"PENALTIES\s+Avg\. Per Game",
        r"TIME OF POSSESSION\s+Avg\. Per Game"
    )

    penalties = "Avg. Per Game " + penalties if penalties else ""

    possession = get_section_by_pattern(
        r"TIME OF POSSESSION\s+Avg\. Per Game",
        r"MISCELLANEOUS\s+3"
    )

    possession = "Avg. Per Game " + possession if possession else ""

    misc = get_section_by_pattern(
        r"MISCELLANEOUS\s+3",
        r"Individual|Participation|Copyright"
    )

    misc = "3" + misc if misc else text

    # ----------------------------
    # Rushing
    # ----------------------------
    team, opp = two_numbers(rushing, r"Yards Gained(?:\s+YDS)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["rush_yards_gained"] = clean_value(team)
    stats["opp_rush_yards_gained"] = clean_value(opp)

    team, opp = two_numbers(rushing, r"Yards Lost(?:\s+YDL)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["rush_yards_lost"] = clean_value(team)
    stats["opp_rush_yards_lost"] = clean_value(opp)

    team, opp = two_numbers(rushing, r"Attempts(?:\s+ATT)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["rush_attempts"] = clean_value(team)
    stats["opp_rush_attempts"] = clean_value(opp)

    team, opp = two_numbers(rushing, r"Average Per Attempt(?:\s+AVG/ATT)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["avg_rush_yards_attempt"] = clean_value(team)
    stats["opp_avg_rush_yards_attempt"] = clean_value(opp)

    team, opp = two_numbers(rushing, r"Avg\. Per Game(?:\s+AVG/G)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["avg_rush_yards_game"] = clean_value(team)
    stats["opp_avg_rush_yards_game"] = clean_value(opp)

    team, opp = two_numbers(rushing, r"Touchdowns(?:\s+TDS)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["rush_tds"] = clean_value(team)
    stats["opp_rush_tds"] = clean_value(opp)

    team, opp = two_numbers(rushing, r"Total(?:\s+TOT)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["rush_yards"] = clean_value(team)
    stats["opp_rush_yards"] = clean_value(opp)

    # ----------------------------
    # Passing
    # ----------------------------
    team, opp = two_ratios(
        passing,
        r"Att-Comp-Int(?:\s+ATT-CMP-INT)?\s+([0-9]+-[0-9]+-[0-9]+)\s+([0-9]+-[0-9]+-[0-9]+)"
    )

    if team and opp:
        stats["pass_attempts"] = clean_value(team.split("-")[0])
        stats["opp_pass_attempts"] = clean_value(opp.split("-")[0])
        stats["interceptions_thrown"] = clean_value(team.split("-")[2])
        stats["interceptions"] = clean_value(opp.split("-")[2])

    team, opp = two_numbers(passing, r"Avg\. Per Attempt(?:\s+AVG/ATT)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["avg_pass_yards_attempt"] = clean_value(team)
    stats["opp_avg_pass_yards_attempt"] = clean_value(opp)

    team, opp = two_numbers(passing, r"Avg\. Per Game(?:\s+AVG/G)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["avg_pass_yards_game"] = clean_value(team)
    stats["opp_avg_pass_yards_game"] = clean_value(opp)

    team, opp = two_numbers(passing, r"Touchdowns(?:\s+TDS)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["pass_tds"] = clean_value(team)
    stats["opp_pass_tds"] = clean_value(opp)

    team, opp = two_numbers(passing, r"Total(?:\s+TOT)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["pass_yards"] = clean_value(team)
    stats["opp_pass_yards"] = clean_value(opp)

    # ----------------------------
    # Total offense
    # ----------------------------
    team, opp = two_numbers(total_offense, r"Total Plays(?:\s+TOT)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["total_offensive_plays"] = clean_value(team)
    stats["total_defensive_plays"] = clean_value(opp)

    team, opp = two_numbers(total_offense, r"Avg\. Per Play(?:\s+AVG/P)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["avg_yards_play"] = clean_value(team)
    stats["opp_avg_yards_play"] = clean_value(opp)

    team, opp = two_numbers(total_offense, r"Avg\. Per Game(?:\s+AVG/G)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["avg_yards_game"] = clean_value(team)
    stats["opp_avg_yards_game"] = clean_value(opp)

    team, opp = two_numbers(total_offense, r"Touchdowns(?:\s+TDS)?\s+([0-9.]+)\s+([0-9.]+)")
    stats["tds_scored"] = clean_value(team)
    stats["opp_tds_scored"] = clean_value(opp)

    # ----------------------------
    # Penalties
    # ----------------------------
    team, opp = two_numbers(
        penalties,
        r"Avg\. Per Game(?:\s+\(YDS\))?(?:\s+AVG/G)?\s+([0-9.]+)\s+([0-9.]+)"
    )

    stats["avg_penalty_yards_game"] = clean_value(team)
    stats["opp_avg_penalty_yards_game"] = clean_value(opp)

    team, opp = two_ratios(penalties, r"Total-Yards(?:\s+TOT/YDS)?\s+([0-9]+-[0-9]+)\s+([0-9]+-[0-9]+)")
    if team and opp:
        stats["total_penalties"] = clean_value(team.split("-")[0])
        stats["opp_total_penalties"] = clean_value(opp.split("-")[0])

    # ----------------------------
    # Time of possession
    # ----------------------------
    team, opp = two_numbers(possession, r"Avg\. Per Game(?:\s+AVG/G)?\s+([0-9]+:[0-9]+)\s+([0-9]+:[0-9]+)")
    stats["time_of_possession_game"] = time_to_minutes(team)
    stats["opp_time_of_possession_game"] = time_to_minutes(opp)

    # ----------------------------
    # Miscellaneous
    # ----------------------------
    team, opp = two_ratios(misc, r"3.*?Down Conversions.*?([0-9]+-[0-9]+).*?([0-9]+-[0-9]+)")
    if team and opp:
        stats["third_down_conv"] = conversion_to_percentage(team)
        stats["opp_third_down_conv"] = conversion_to_percentage(opp)

    team, opp = two_ratios(misc, r"4.*?Down Conversions.*?([0-9]+-[0-9]+).*?([0-9]+-[0-9]+)")
    if team and opp:
        stats["fourth_down_conv"] = conversion_to_percentage(team)
        stats["opp_fourth_down_conv"] = conversion_to_percentage(opp)

    team, opp = two_ratios(misc, r"Sacks-Yards Lost(?:\s+SCK YDS)?\s+([0-9]+-[0-9]+)\s+([0-9]+-[0-9]+)")
    if team and opp:
        stats["sacks"] = clean_value(team.split("-")[0])
        stats["sacks_taken"] = clean_value(opp.split("-")[0])

    return stats


def parse_stat_file(file_path_or_url):
    """
    Chooses the correct parser based on whether the input is:
    - URL
    - XML file
    - HTML file
    - PDF file
    """
    file_path_or_url = str(file_path_or_url)

    if file_path_or_url.startswith("http://") or file_path_or_url.startswith("https://"):
        return parse_url_file(file_path_or_url)

    file_path = Path(file_path_or_url)
    suffix = file_path.suffix.lower()

    if suffix == ".xml":
        return parse_xml_file(file_path)

    if suffix in [".html", ".htm"]:
        return parse_html_file(file_path)

    if suffix == ".pdf":
        return parse_pdf_file(file_path)

    raise ValueError(f"Unsupported file type: {file_path}")


def find_stat_value(excel_header, parsed_stats):
    """
    Uses STAT_ALIASES to find which scraped value belongs under
    a specific Excel column.
    """
    possible_names = STAT_ALIASES.get(excel_header, [])

    for name in possible_names:
        cleaned_name = clean_key(name)

        if cleaned_name in parsed_stats:
            return parsed_stats[cleaned_name]

    return None


def update_workbook(team, year, parsed_stats):
    """
    Opens the correct team Excel file, finds the requested year row,
    and fills in any stats that were found in the input file.
    """
    excel_path = EXCEL_DIR / TEAM_FILES[team]

    if not excel_path.exists():
        raise FileNotFoundError(f"Could not find Excel file: {excel_path}")

    workbook = load_workbook(excel_path)
    sheet = workbook["Sheet1"]

    # Read the Excel header row so we know which stat belongs in each column.
    header_to_col = {}

    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value

        if header:
            header_to_col[header] = col

    # Find the row matching the selected year.
    target_row = None

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == year:
            target_row = row
            break

    if target_row is None:
        raise ValueError(f"Year {year} not found in {excel_path.name}")

    updated_count = 0

    # Fill in matching stats.
    for header, col in header_to_col.items():
        if header == "Year":
            continue

        value = find_stat_value(header, parsed_stats)

        if value is not None:
            sheet.cell(row=target_row, column=col).value = value
            updated_count += 1

    # Save to output folder so the original Excel file stays safe.
    output_path = OUTPUT_DIR / excel_path.name
    workbook.save(output_path)

    print()
    print(f"Updated {updated_count} stats for {team.upper()} {year}")
    print(f"Saved file to: {output_path}")


def main():
    """
    Main program flow:
    1. Ask user which team/year/file to use
    2. Parse the XML/HTML file
    3. Update the matching Excel workbook
    """
    team = input("Team name csup/mesa/mines/western: ").strip().lower()
    year = int(input("Year: ").strip())
    input_source = input("Input filename or URL: ").strip()

    if team not in TEAM_FILES:
        raise ValueError("Invalid team name. Use csup, mesa, mines, or western.")

    if input_source.startswith("http://") or input_source.startswith("https://"):
        parsed_stats = parse_stat_file(input_source)
    else:
        input_path = INPUT_DIR / input_source

        if not input_path.exists():
            raise FileNotFoundError(f"Could not find input file: {input_path}")

        parsed_stats = parse_stat_file(input_path)

    print()
    print("Parsed stats found:")
    for key, value in parsed_stats.items():
        print(f"{key}: {value}")

    update_workbook(team, year, parsed_stats)


if __name__ == "__main__":
    main()