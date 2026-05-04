from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas

import scraper


def test_clean_key_basic_formatting():
    assert scraper.clean_key("Rush Yards") == "rush_yards"
    assert scraper.clean_key("Opp. Avg Yards/Game") == "opp_avg_yards_game"
    assert scraper.clean_key("3rd Down Conv.") == "3rd_down_conv"


def test_clean_value_converts_numbers():
    assert scraper.clean_value("1850") == 1850
    assert scraper.clean_value("4.5") == 4.5
    assert scraper.clean_value("55%") == 0.55


def test_clean_value_handles_empty_values():
    assert scraper.clean_value("") is None
    assert scraper.clean_value("   ") is None
    assert scraper.clean_value(None) is None


def test_clean_value_keeps_text_when_not_numeric():
    assert scraper.clean_value("12:34") == "12:34"
    assert scraper.clean_value("N/A") == "N/A"


def test_parse_xml_file_reads_tags(tmp_path):
    xml_file = tmp_path / "sample.xml"

    xml_file.write_text(
        """
        <football_stats>
            <rush_yards>1850</rush_yards>
            <rush_attempts>410</rush_attempts>
            <rush_tds>22</rush_tds>
        </football_stats>
        """
    )

    stats = scraper.parse_xml_file(xml_file)

    assert stats["rush_yards"] == 1850
    assert stats["rush_attempts"] == 410
    assert stats["rush_tds"] == 22


def test_parse_xml_file_reads_attributes(tmp_path):
    xml_file = tmp_path / "sample.xml"

    xml_file.write_text(
        """
        <football_stats>
            <stat name="rush_yards" value="1850" />
        </football_stats>
        """
    )

    stats = scraper.parse_xml_file(xml_file)

    assert stats["name"] == "rush_yards"
    assert stats["value"] == 1850


def test_parse_html_file_reads_table_rows(tmp_path):
    html_file = tmp_path / "sample.html"

    html_file.write_text(
        """
        <html>
            <body>
                <table>
                    <tr><th>rush_yards</th><td>1600</td></tr>
                    <tr><th>rush_attempts</th><td>388</td></tr>
                    <tr><th>rush_tds</th><td>19</td></tr>
                </table>
            </body>
        </html>
        """
    )

    stats = scraper.parse_html_file(html_file)

    assert stats["rush_yards"] == 1600
    assert stats["rush_attempts"] == 388
    assert stats["rush_tds"] == 19


def test_parse_stat_file_rejects_bad_file_type(tmp_path):
    bad_file = tmp_path / "sample.txt"
    bad_file.write_text("not valid")

    with pytest.raises(ValueError):
        scraper.parse_stat_file(bad_file)


def test_find_stat_value_uses_aliases():
    parsed_stats = {
        "rush_yards": 1850,
        "rush_attempts": 410,
    }

    assert scraper.find_stat_value("Rush Yards", parsed_stats) == 1850
    assert scraper.find_stat_value("Rush Attempts", parsed_stats) == 410


def test_find_stat_value_returns_none_when_missing():
    parsed_stats = {
        "rush_yards": 1850,
    }

    assert scraper.find_stat_value("Pass TDs", parsed_stats) is None


def test_update_workbook_writes_to_temp_output_only(tmp_path, monkeypatch):
    fake_excel_dir = tmp_path / "excel_files"
    fake_output_dir = tmp_path / "output"

    fake_excel_dir.mkdir()
    fake_output_dir.mkdir()

    workbook_path = fake_excel_dir / "mines_football_stats.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    sheet["A1"] = "Year"
    sheet["B1"] = "Rush Yards"
    sheet["C1"] = "Rush Attempts"
    sheet["D1"] = "Rush TDs"

    sheet["A2"] = 2015
    sheet["A3"] = 2016

    workbook.save(workbook_path)

    monkeypatch.setattr(scraper, "EXCEL_DIR", fake_excel_dir)
    monkeypatch.setattr(scraper, "OUTPUT_DIR", fake_output_dir)

    parsed_stats = {
        "rush_yards": 1850,
        "rush_attempts": 410,
        "rush_tds": 22,
    }

    scraper.update_workbook("mines", 2015, parsed_stats)

    output_file = fake_output_dir / "mines_football_stats.xlsx"
    assert output_file.exists()

    updated_workbook = load_workbook(output_file)
    updated_sheet = updated_workbook["Sheet1"]

    assert updated_sheet["A2"].value == 2015
    assert updated_sheet["B2"].value == 1850
    assert updated_sheet["C2"].value == 410
    assert updated_sheet["D2"].value == 22


def test_update_workbook_raises_error_for_missing_year(tmp_path, monkeypatch):
    fake_excel_dir = tmp_path / "excel_files"
    fake_output_dir = tmp_path / "output"

    fake_excel_dir.mkdir()
    fake_output_dir.mkdir()

    workbook_path = fake_excel_dir / "mines_football_stats.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    sheet["A1"] = "Year"
    sheet["B1"] = "Rush Yards"
    sheet["A2"] = 2015

    workbook.save(workbook_path)

    monkeypatch.setattr(scraper, "EXCEL_DIR", fake_excel_dir)
    monkeypatch.setattr(scraper, "OUTPUT_DIR", fake_output_dir)

    with pytest.raises(ValueError):
        scraper.update_workbook("mines", 2020, {"rush_yards": 2000})


def test_parse_stat_file_accepts_url(monkeypatch):
    def fake_parse_url_file(url):
        return {"rush_yards": 1850}

    monkeypatch.setattr(scraper, "parse_url_file", fake_parse_url_file)

    stats = scraper.parse_stat_file("https://example.com/stats")

    assert stats["rush_yards"] == 1850


def test_parse_pdf_file_extracts_team_stats_2015_style(tmp_path):
    pdf_file = tmp_path / "sample.pdf"

    c = canvas.Canvas(str(pdf_file))
    c.drawString(50, 750, "Team Statistics CSUP OPP")
    c.drawString(50, 730, "RUSHING YARDAGE 4074 1324")
    c.drawString(50, 710, " Yards gained 4459 1822")
    c.drawString(50, 690, " Yards lost 385 498")
    c.drawString(50, 670, " Attempts 627 460")
    c.drawString(50, 650, " Avg Per Attempt 6.5 2.9")
    c.drawString(50, 630, " Avg Per Game 291.0 94.6")
    c.drawString(50, 610, " TDs Passing 47 12")
    c.drawString(50, 590, "PASSING YARDAGE 1750 2999")
    c.drawString(50, 570, " Att-Comp-Int 229-124-12 533-306-24")
    c.drawString(50, 550, " Avg Per Attempt 7.64 5.63")
    c.drawString(50, 530, " Avg Per Game 125.00 214.21")
    c.drawString(50, 510, " Touchdowns 16 16")
    c.drawString(50, 490, "TOTAL OFFENSE 5824 4323")
    c.drawString(50, 470, " Total Plays 856 993")
    c.drawString(50, 450, " Avg Per Play 6.8 4.4")
    c.drawString(50, 430, " Avg Per Game 416.0 308.8")
    c.drawString(50, 410, "PENALTIES-YARDS 95-918 89-850")
    c.drawString(50, 390, " Avg Per Game 65.57 60.71")
    c.drawString(50, 370, "TIME OF POSSESSION 06:52:29 07:07:31")
    c.drawString(50, 350, " Per Game 29:28 30:32")
    c.drawString(50, 330, "3RD-DOWN CONVERSIONS 55-167 76-238")
    c.drawString(50, 310, "4TH-DOWN CONVERSIONS 3-15 9-31")
    c.drawString(50, 290, "SACKS BY-YARDS 42-300 19-128")
    c.drawString(50, 270, "TOUCHDOWNS SCORED 68 28")
    c.save()

    stats = scraper.parse_pdf_file(pdf_file)

    assert stats["rush_yards"] == 4074
    assert stats["opp_rush_yards"] == 1324
    assert stats["rush_yards_gained"] == 4459
    assert stats["rush_yards_lost"] == 385
    assert stats["rush_attempts"] == 627
    assert stats["avg_rush_yards_attempt"] == 6.5
    assert stats["avg_rush_yards_game"] == 291.0
    assert stats["rush_tds"] == 47

    assert stats["pass_attempts"] == 229
    assert stats["interceptions_thrown"] == 12
    assert stats["pass_tds"] == 16
    assert stats["opp_pass_attempts"] == 533
    assert stats["interceptions"] == 24

    assert stats["total_offensive_plays"] == 856
    assert stats["total_defensive_plays"] == 993
    assert stats["avg_yards_play"] == 6.8
    assert stats["opp_avg_yards_play"] == 4.4

    assert stats["total_penalties"] == 95
    assert stats["opp_total_penalties"] == 89
    assert stats["third_down_conv"] == round(55 / 167, 4)
    assert stats["fourth_down_conv"] == round(3 / 15, 4)
    assert stats["sacks"] == 42
    assert stats["sacks_taken"] == 19
    assert stats["tds_scored"] == 68
    assert stats["opp_tds_scored"] == 28


def test_parse_stat_file_accepts_team_stats_pdf(tmp_path):
    pdf_file = tmp_path / "sample.pdf"

    c = canvas.Canvas(str(pdf_file))
    c.drawString(50, 750, "Team Statistics CSUP OPP")
    c.drawString(50, 730, "RUSHING YARDAGE 2154 925")
    c.drawString(50, 710, " Yards gained 2352 1228")
    c.drawString(50, 690, " Yards lost 198 303")
    c.drawString(50, 670, " Attempts 410 334")
    c.drawString(50, 650, " TDs Passing 24 7")
    c.drawString(50, 630, "PASSING YARDAGE 2030 2440")
    c.drawString(50, 610, " Att-Comp-Int 300-178-12 391-245-14")
    c.drawString(50, 590, " Touchdowns 17 16")
    c.save()

    stats = scraper.parse_stat_file(pdf_file)

    assert stats["rush_yards"] == 2154
    assert stats["opp_rush_yards"] == 925
    assert stats["rush_attempts"] == 410
    assert stats["opp_rush_attempts"] == 334
    assert stats["rush_tds"] == 24
    assert stats["opp_rush_tds"] == 7
    assert stats["pass_attempts"] == 300
    assert stats["opp_pass_attempts"] == 391
    assert stats["interceptions_thrown"] == 12
    assert stats["interceptions"] == 14
    assert stats["pass_tds"] == 17
    assert stats["opp_pass_tds"] == 16


def test_parse_pdf_file_ignores_schedule_values_after_stats(tmp_path):
    pdf_file = tmp_path / "sample.pdf"

    c = canvas.Canvas(str(pdf_file))
    c.drawString(50, 750, "Team Statistics CUFB OPP Date Opponent Score Record Time Attend")
    c.drawString(50, 730, "RUSHING YARDAGE 2171 1327")
    c.drawString(50, 710, "Yards gained 2404 1620 *10/14/2017 at Fort Lewis L 24-35 5-2 , 5-1 2:55 2,123")
    c.drawString(50, 690, "Yards lost 233 293 *10/21/2017 vs Black Hills State W 49-0 6-2 , 6-1 2:28 6,023")
    c.drawString(50, 670, "Attempts 429 419 *10/28/2017 vs Western State W 40-7 7-2 , 7-1 2:32 5,324")
    c.drawString(50, 650, "Avg Per Attempt 5.1 3.2 *11/4/2017 at Dixie State W 31-10 8-2 , 8-1 2:39 2,580")
    c.drawString(50, 630, "Avg Per Game 180.9 110.6 *11/11/2017 at Chadron State W 28-6 9-2 , 9-1 2:56 2,630")
    c.drawString(50, 610, "TDs Passing 29 7 11/18/2017 at Minnesota State L 13-16 9-3 2:48 1,275")
    c.save()

    stats = scraper.parse_pdf_file(pdf_file)

    assert stats["rush_yards"] == 2171
    assert stats["opp_rush_yards"] == 1327
    assert stats["rush_yards_gained"] == 2404
    assert stats["opp_rush_yards_gained"] == 1620
    assert stats["rush_yards_lost"] == 233
    assert stats["opp_rush_yards_lost"] == 293
    assert stats["rush_attempts"] == 429
    assert stats["opp_rush_attempts"] == 419
    assert stats["avg_rush_yards_attempt"] == 5.1
    assert stats["opp_avg_rush_yards_attempt"] == 3.2
    assert stats["avg_rush_yards_game"] == 180.9
    assert stats["opp_avg_rush_yards_game"] == 110.6
    assert stats["rush_tds"] == 29
    assert stats["opp_rush_tds"] == 7


def test_parse_url_file_extracts_csu_clean_page_layout(monkeypatch):
    class FakeResponse:
        text = """
        <html>
            <body>
                <p>Team Stats (11-2, 9-1)</p>
                <p>Overall Team Statistics Statistic CSU-PUEBLO Opponents OPP</p>

                <p>Rushing</p>
                <p>Yards Gained YDS 3615 1584</p>
                <p>Yards Lost YDL 355 503</p>
                <p>Attempts ATT 598 507</p>
                <p>Average Per Attempt AVG/ATT 5.5 2.1</p>
                <p>Avg. Per Game AVG/G 250.8 83.2</p>
                <p>Touchdowns TDS 40 8</p>
                <p>Total TOT 3260 1081</p>

                <p>Passing</p>
                <p>Att-Comp-Int ATT-CMP-INT 323-170-6 403-196-28</p>
                <p>Avg. Per Attempt AVG/ATT 6.40 5.12</p>
                <p>Avg. Per Game AVG/G 159.00 158.62</p>
                <p>Touchdowns TDS 13 11</p>
                <p>Total TOT 2067 2062</p>

                <p>Total Offense</p>
                <p>Total Plays TOT 921 910</p>
                <p>Avg. Per Play AVG/P 5.8 3.5</p>
                <p>Avg. Per Game AVG/G 409.8 241.8</p>
                <p>Touchdowns TDS 59 22</p>

                <p>Penalties</p>
                <p>Avg. Per Game (YDS) AVG/G 70.46 46.15</p>
                <p>Total-Yards TOT/YDS 89-916 62-600</p>

                <p>Time Of Possession</p>
                <p>Avg. Per Game AVG/G 30:01 29:52</p>

                <p>Miscellaneous</p>
                <p>3^{rd} Down Conversions 3^{rd} % 60-201 (29.85%) 57-214 (26.64%)</p>
                <p>4^{th} Down Conversions 4^{th} % 13-25 (52.00%) 8-23 (34.78%)</p>
                <p>Sacks-Yards Lost SCK YDS 52-307 25-154</p>
            </body>
        </html>
        """

        def raise_for_status(self):
            pass

    def fake_get(url, timeout):
        return FakeResponse()

    monkeypatch.setattr(scraper.requests, "get", fake_get)

    stats = scraper.parse_url_file("https://example.com/stats")

    assert stats["rush_yards"] == 3260
    assert stats["opp_rush_yards"] == 1081
    assert stats["rush_yards_gained"] == 3615
    assert stats["opp_rush_yards_gained"] == 1584
    assert stats["rush_yards_lost"] == 355
    assert stats["opp_rush_yards_lost"] == 503
    assert stats["rush_attempts"] == 598
    assert stats["opp_rush_attempts"] == 507
    assert stats["rush_tds"] == 40
    assert stats["opp_rush_tds"] == 8

    assert stats["pass_attempts"] == 323
    assert stats["opp_pass_attempts"] == 403
    assert stats["interceptions_thrown"] == 6
    assert stats["interceptions"] == 28
    assert stats["pass_tds"] == 13
    assert stats["opp_pass_tds"] == 11

    assert stats["total_offensive_plays"] == 921
    assert stats["total_defensive_plays"] == 910
    assert stats["avg_yards_play"] == 5.8
    assert stats["opp_avg_yards_play"] == 3.5
    assert stats["tds_scored"] == 59
    assert stats["opp_tds_scored"] == 22

    assert stats["total_penalties"] == 89
    assert stats["opp_total_penalties"] == 62
    assert stats["time_of_possession_game"] == 30.02
    assert stats["opp_time_of_possession_game"] == 29.87

    assert stats["third_down_conv"] == round(60 / 201, 4)
    assert stats["opp_third_down_conv"] == round(57 / 214, 4)
    assert stats["fourth_down_conv"] == round(13 / 25, 4)
    assert stats["opp_fourth_down_conv"] == round(8 / 23, 4)

    assert stats["sacks"] == 52
    assert stats["sacks_taken"] == 25